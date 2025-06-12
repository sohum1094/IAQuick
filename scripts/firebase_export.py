import argparse
import os
from datetime import datetime
from copy import copy

import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl


def init_firebase(cred_path=None):
    if not firebase_admin._apps:
        if cred_path:
            cred = credentials.Certificate(cred_path)
        elif os.environ.get("GOOGLE_APPLICATION_CREDENTIALS"):
            cred = credentials.Certificate(os.environ["GOOGLE_APPLICATION_CREDENTIALS"])
        else:
            cred = credentials.ApplicationDefault()
        firebase_admin.initialize_app(cred)


def fetch_survey_data(survey_id):
    db = firestore.client()
    survey_ref = db.collection('surveys').document(survey_id)
    doc = survey_ref.get()
    if not doc.exists:
        raise ValueError(f"Survey {survey_id} not found")
    header = doc.to_dict()

    rooms = [d.to_dict() for d in survey_ref.collection('room_readings').get()]
    visuals = [d.to_dict() for d in survey_ref.collection('visualAssessments').get()]
    return header, rooms, visuals


def parse_date(value):
    if isinstance(value, datetime):
        return value
    if hasattr(value, 'to_datetime'):
        return value.to_datetime()
    if isinstance(value, str):
        try:
            return datetime.fromisoformat(value)
        except Exception:
            pass
    return datetime.utcnow()


def copy_row_style(ws, template_row_idx, target_row_idx):
    template_cells = [cell for cell in ws[template_row_idx]]
    ws.insert_rows(target_row_idx)
    for col_idx, tmpl in enumerate(template_cells, start=1):
        new_cell = ws.cell(row=target_row_idx, column=col_idx)
        new_cell._style = copy(tmpl._style)
        if tmpl.has_style and tmpl.font is not None:
            new_cell.font = copy(tmpl.font)
        if tmpl.fill is not None:
            new_cell.fill = copy(tmpl.fill)
        if tmpl.border is not None:
            new_cell.border = copy(tmpl.border)
        if tmpl.number_format:
            new_cell.number_format = tmpl.number_format
        if tmpl.alignment:
            new_cell.alignment = copy(tmpl.alignment)
        if tmpl.protection:
            new_cell.protection = copy(tmpl.protection)


def export_iaq(header, rooms, output_dir, template_path='assets/IAQ_template_v2.xlsx'):
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Data for Print']

    site_name = header.get('siteName', '[School Name]')
    date = parse_date(header.get('date') or header.get('surveyDate'))
    occupancy = header.get('occupancyType') or header.get('occupancyStatus', '')

    ws['A1'] = f"{site_name} Indoor Air Quality Measurements"
    ws['A2'] = date.strftime('%m/%d/%Y')
    ws['A3'] = occupancy

    # clear existing template rows beyond the header
    if ws.max_row > 5:
        ws.delete_rows(5, ws.max_row - 4)

    template_idx = 5
    rooms = sorted(rooms, key=lambda r: parse_date(r.get('timestamp', '')))
    first_outdoor = next((r for r in rooms if r.get('isOutdoor')), None)
    if first_outdoor:
        rooms.remove(first_outdoor)
        rooms.insert(0, first_outdoor)
    stats = {'temp': [], 'rh': [], 'co2': [], 'pm25': []}
    for i, r in enumerate(rooms, start=0):
        row_idx = template_idx + i
        copy_row_style(ws, template_idx, row_idx)
        ws.cell(row=row_idx, column=1, value=r.get('building'))
        ws.cell(row=row_idx, column=2, value=r.get('floorNumber'))
        ws.cell(row=row_idx, column=3, value=r.get('roomNumber'))
        ws.cell(row=row_idx, column=4, value=r.get('primaryUse') or r.get('primaryRoomUse'))
        t = r.get('temperature') or r.get('temperatureF')
        rh = r.get('relativeHumidity') or r.get('relativeHumidityPct')
        co2 = r.get('co2') or r.get('co2ppm')
        pm25 = r.get('pm25') or r.get('pm25mgm3')
        ws.cell(row=row_idx, column=5, value=t)
        ws.cell(row=row_idx, column=6, value=rh)
        ws.cell(row=row_idx, column=7, value=co2)
        ws.cell(row=row_idx, column=8, value=pm25)
        if t is not None:
            stats['temp'].append(t)
        if rh is not None:
            stats['rh'].append(rh)
        if co2 is not None:
            stats['co2'].append(co2)
        if pm25 is not None:
            stats['pm25'].append(pm25)

    # min & max sheet
    mm = wb.create_sheet('Min&Max') if 'Min&Max' not in wb.sheetnames else wb['Min&Max']
    mm['A1'] = 'Parameter'
    mm['B1'] = 'Min'
    mm['C1'] = 'Max'
    def write_stat(row, label, values):
        mm.cell(row=row, column=1, value=label)
        if values:
            mm.cell(row=row, column=2, value=min(values))
            mm.cell(row=row, column=3, value=max(values))
    write_stat(2, 'Temperature (F)', stats['temp'])
    write_stat(3, 'Relative Humidity (%)', stats['rh'])
    write_stat(4, 'CO2 (ppm)', stats['co2'])
    write_stat(5, 'PM2.5 (ug/m3)', stats['pm25'])

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f"{site_name.replace(' ', '_')}_IAQ_{date.strftime('%Y%m%d')}.xlsx")
    wb.save(out_path)
    return out_path


def export_visual(header, visuals, output_dir, template_path='assets/Visual_template.xlsx'):
    wb = openpyxl.load_workbook(template_path)
    entry = wb['Entry Sheet']
    va = wb['VA for Print']

    site_name = header.get('siteName', '[Site]')
    date = parse_date(header.get('date') or header.get('surveyDate'))
    occupancy = header.get('occupancyType') or header.get('occupancyStatus', '')

    entry['A2'] = occupancy
    entry['B2'] = date
    entry['D2'] = site_name
    va['A2'] = date
    va['A3'] = occupancy

    if va.max_row > 5:
        va.delete_rows(5, va.max_row - 4)
    template_idx = 5
    for i, v in enumerate(visuals, start=0):
        row_idx = template_idx + i
        copy_row_style(va, template_idx, row_idx)
        va.cell(row=row_idx, column=1, value=v.get('building'))
        va.cell(row=row_idx, column=2, value=v.get('floorNumber'))
        va.cell(row=row_idx, column=3, value=v.get('roomNumber'))
        va.cell(row=row_idx, column=4, value=v.get('primaryRoomUse') or v.get('primaryUse'))
        va.cell(row=row_idx, column=5, value=v.get('notes'))

    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f"{site_name.replace(' ', '_')}_Visual_Assessment_{date.strftime('%Y%m%d')}.xlsx")
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Export survey data from Firebase to Excel")
    parser.add_argument('surveyId', help='Firestore survey document ID')
    parser.add_argument('--credentials', help='Path to service account JSON')
    parser.add_argument('--output', default='output', help='Base output directory')
    args = parser.parse_args()

    init_firebase(args.credentials)
    header, rooms, visuals = fetch_survey_data(args.surveyId)
    site = header.get('siteName', 'Site').replace(' ', '_')
    out_dir = os.path.join(args.output, site)
    if rooms:
        export_iaq(header, rooms, out_dir)
    if visuals:
        export_visual(header, visuals, out_dir)


if __name__ == '__main__':
    main()

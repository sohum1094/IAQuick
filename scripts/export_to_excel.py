import openpyxl
from openpyxl import Workbook
from datetime import datetime
import sys


def create_iaq_report(data, output_path, template_path='assets/IAQ_template_v2.xlsx'):
    """Create IAQ Excel file replicating the template layout.

    Parameters
    ----------
    data : list of dict
        Each dict should contain keys: building, floor, room, primary_use,
        temperature, humidity, co2, pm25
    output_path : str
        Where to save the resulting XLSX file.
    template_path : str
        Path to the template Excel file.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active  # 'Data for Print'

    # Fill header information
    ws['A1'] = data[0].get('site_name', '[School Name]') if data else '[School Name]'
    ws['A2'] = data[0].get('date', datetime.today()).strftime('%m/%d/%Y') if data else datetime.today().strftime('%m/%d/%Y')
    ws['A3'] = data[0].get('occupancy', '[Occupancy type]') if data else '[Occupancy type]'

    start_row = 5
    for idx, entry in enumerate(data, start=0):
        row = start_row + idx
        ws.cell(row=row, column=1, value=entry.get('building'))
        ws.cell(row=row, column=2, value=entry.get('floor'))
        ws.cell(row=row, column=3, value=entry.get('room'))
        ws.cell(row=row, column=4, value=entry.get('primary_use'))
        ws.cell(row=row, column=5, value=entry.get('temperature'))
        ws.cell(row=row, column=6, value=entry.get('humidity'))
        ws.cell(row=row, column=7, value=entry.get('co2'))
        ws.cell(row=row, column=8, value=entry.get('pm25'))

    wb.save(output_path)


if __name__ == '__main__':
    # Example usage with dummy data
    example_data = [
        {
            'site_name': 'Example School',
            'date': datetime(2024, 1, 1),
            'occupancy': 'Occupied',
            'building': 'A',
            'floor': 1,
            'room': '101',
            'primary_use': 'Classroom',
            'temperature': 72,
            'humidity': 40,
            'co2': 750,
            'pm25': 12,
        }
    ]
    output = sys.argv[1] if len(sys.argv) > 1 else 'output.xlsx'
    create_iaq_report(example_data, output)
